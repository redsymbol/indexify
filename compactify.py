#!/usr/bin/env python3

import os
import re
import argparse
import sys

# error exit codes
ERR_CANNOT_OVERWRITE_FILE = 1

from warnings import warn
from collections import (
    namedtuple,
    defaultdict,
    OrderedDict,
    )
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell

def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('input_xls')
    parser.add_argument('output_xls', nargs='?', default=None)
    parser.add_argument('-O', '--overwrite', action='store_true', default=False,
                        help='Overwrite output file, if it exists')
    return parser.parse_args()

RawIndexEntry = namedtuple('RawIndexEntry', 'heading pages see_also')
RawLocator = namedtuple('RawLocator', 'pages see_also')

class Locator:
    def __init__(self):
        self.pages = set()
        self.see_also = set()
    def render(self):
        pages = render_pages(self.pages)
        see_also = ", ".join(sorted(self.see_also))
        if pages == "":
            return "see " + see_also
        if see_also == "":
            return pages
        return pages + ", see also " + see_also

class Index:
    def __init__(self):
        # Mapping index heading to locator value (i.e., locator rendered to string)
        self.entries = OrderedDict()
    @classmethod
    def from_raw_index(cls, raw_index):
        index = cls()
        for heading, locator in sorted(raw_index.items(), key=heading_sort_key):
            index.entries[heading] = locator.render()
        return index
    def write_spreadsheet(self, dest: str, overwrite: bool = False):
        wb = Workbook()
        ws = wb.active
        def mk_cell(value):
            cell = Cell(ws)
            cell.number_format = '@' # Ensure text cell format
            cell.set_explicit_value(value)
            return cell
        for heading, locator_value in self.entries.items():
            ws.append([
                mk_cell(heading),
                mk_cell(locator_value),
            ])
        # openpyxl.Workbook.save silently overwrites its target, if
        # the file already exists.  This clobber check is a slight
        # race condition, but can't be helped without modifying that
        # save method itself.
        if (not overwrite) and os.path.exists(dest):
            raise FileExistsError(dest)
        wb.save(dest)
    def pretty_print(self):
        for heading, locator_value in self.entries.items():
            print("{}: {}".format(heading, locator_value))

def render_pages(page_set):
    if len(page_set) == 0:
        return ""
    intervals = []
    pages = sorted(page_set)
    last = pages.pop(0)
    start = last
    page = last
    while pages:
        page = pages.pop(0)
        if last + 1 == page:
            last = page
            continue
        intervals.append((start, last))
        start = page
        last = page
    intervals.append((start, page))
    substrings = []
    for start, end in intervals:
        if start == end:
            substrings.append(str(start))
        else:
            substrings.append("{}-{}".format(start, end))
    return ", ".join(substrings)

def heading_sort_key(pair):
    heading = pair[0]
    key = heading.lower()
    if len(key) > 0 and key[0] in '"“':
        return key[1:]
    return key

def raw_index_from_entries(entries):
    raw_index = defaultdict(Locator)
    for entry in entries:
        locator = raw_index[entry.heading]
        locator.pages |= entry.pages
        locator.see_also |= entry.see_also
    return raw_index

def raw_index_entries(input_xls):
    wb = load_workbook(filename=input_xls, read_only=True)
    ws = wb.active
    for row in ws.rows:
        heading, pagespec = row[0].value, str(row[1].value)
        try:
            raw_locator = pagespec2rawlocator(pagespec)
        except ValueError:
            warn("Cannot parse row: {}: {}".format(heading, pagespec))
            continue
        yield RawIndexEntry(
            heading,
            raw_locator.pages,
            raw_locator.see_also)

def pagespec2rawlocator(pages_spec):
    first, *remain = pages_spec.split(',')
    first = first.strip()
    see_also = set()
    if re.search(r'^\d+$', first):
        pages = {int(first)}
    elif re.search(r'^\d+-\d+$', first):
        start_s, end_s = first.split('-')
        pages = pagerange(start_s, end_s)
    elif re.search(r'^see ', first):
        pages = set()
        see_also = {first[4:]}
    else:
        raise ValueError(first)
    for spec in remain:
        rawlocator = pagespec2rawlocator(spec)
        pages |= rawlocator.pages
        see_also |= rawlocator.see_also
    return RawLocator(pages, see_also)

def pagerange(start_s, end_s):
    offset = len(start_s) - len(end_s)
    start_i = int(start_s)
    real_end_s = start_s[:offset] + end_s
    end_i = int(real_end_s)
    return set(range(start_i, end_i + 1))

if __name__ == '__main__':
    args = get_args()
    entries = raw_index_entries(args.input_xls)
    raw_index = raw_index_from_entries(entries)
    index = Index.from_raw_index(raw_index)
    if args.output_xls is None:
        index.pretty_print()
    else:
        try:
            index.write_spreadsheet(args.output_xls, args.overwrite)
        except FileExistsError:
            sys.stderr.write('Output file "{}" already exists. Refusing to overwrite. (Use --overwrite to force.)\n'.format(args.output_xls))
            sys.exit(ERR_CANNOT_OVERWRITE_FILE)
            
